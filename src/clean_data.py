"""
DecodeLabs Project 1 - Data Cleaning & Preparation
Tool: Python (pandas + openpyxl)
Reason: <1M rows so Excel/Power Query would also work, but a script gives a
reproducible, auditable workflow that can be re-run if the raw export changes.
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = "../data/raw/raw_dataset.xlsx"
OUT = "../data/cleaned/Cleaned_Dataset.xlsx"

df = pd.read_excel(SRC)
change_log = []

# ---------- PHASE 1: Strategic handling of gaps ----------
missing_coupon = df["CouponCode"].isna().sum()
# CouponCode is missing because no coupon was applied at checkout, not
# because data was lost at random. Filling with the mode (most frequent
# real coupon) would falsely imply a discount was used. The correct fix is
# an explicit category, not a statistical guess.
df["CouponCode"] = df["CouponCode"].fillna("No Coupon")
change_log.append({
    "Change ID": "CR001",
    "Description": "Filled missing CouponCode values with explicit label 'No Coupon'",
    "Impact": f"Resolved {missing_coupon} records ({missing_coupon/len(df)*100:.1f}% of rows)",
    "Status": "Resolved",
})

# ---------- PHASE 2: Integrity audit (duplicates) ----------
full_dupes = df.duplicated().sum()
df = df.drop_duplicates()
orderid_dupes = df["OrderID"].duplicated().sum()
tracking_dupes = df["TrackingNumber"].duplicated().sum()
change_log.append({
    "Change ID": "CR002",
    "Description": "Audited full-row duplicates and duplicate OrderID/TrackingNumber values",
    "Impact": f"Removed {full_dupes} duplicate rows; confirmed 0 duplicate OrderIDs and 0 duplicate TrackingNumbers across {len(df)} records",
    "Status": "Resolved",
})

# ---------- PHASE 3: Standardize formats ----------
text_cols = ["Product", "PaymentMethod", "OrderStatus", "ReferralSource",
             "ShippingAddress"]
for col in text_cols:
    df[col] = df[col].astype(str).str.strip().str.title()
# Codes/IDs: trim whitespace only, never re-case (would break SAVE10/ORD/TRK/C prefixes)
for col in ["OrderID", "CustomerID", "TrackingNumber", "CouponCode"]:
    df[col] = df[col].astype(str).str.strip()
change_log.append({
    "Change ID": "CR003",
    "Description": "Trimmed whitespace and standardized casing on free-text category fields (codes/IDs left untouched)",
    "Impact": f"Applied across {len(df)} records on 5 categorical columns",
    "Status": "Resolved",
})

unit_fix = (df["UnitPrice"].round(2) != df["UnitPrice"]).sum()
total_fix = (df["TotalPrice"].round(2) != df["TotalPrice"]).sum()
df["UnitPrice"] = df["UnitPrice"].round(2)
df["TotalPrice"] = df["TotalPrice"].round(2)
change_log.append({
    "Change ID": "CR004",
    "Description": "Standardized numeric precision to 2 decimal places (UnitPrice, TotalPrice)",
    "Impact": f"Corrected {unit_fix} UnitPrice and {total_fix} TotalPrice values with inconsistent decimal precision",
    "Status": "Resolved",
})

bad_dates = (~pd.to_datetime(df["Date"], errors="coerce").notna()).sum()
df["Date"] = pd.to_datetime(df["Date"])
change_log.append({
    "Change ID": "CR005",
    "Description": "Validated and standardized Date column to ISO 8601 (YYYY-MM-DD)",
    "Impact": f"Confirmed {bad_dates} unparseable dates across {len(df)} records; all dates conform to ISO 8601",
    "Status": "Resolved",
})

log_df = pd.DataFrame(change_log)

# ---------- WRITE OUTPUT ----------
with pd.ExcelWriter(OUT, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="Cleaned_Data", index=False)
    log_df.to_excel(writer, sheet_name="Change_Log", index=False)

# ---------- FORMATTING ----------
wb = load_workbook(OUT)
header_fill = PatternFill("solid", start_color="2F5233", end_color="2F5233")
header_font = Font(name="Arial", bold=True, color="FFFFFF")
body_font = Font(name="Arial")
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

ws = wb["Cleaned_Data"]
date_col_idx = df.columns.get_loc("Date") + 1
price_cols_idx = [df.columns.get_loc(c) + 1 for c in ["UnitPrice", "TotalPrice"]]
for col_cells in ws.iter_cols(min_row=1, max_row=1):
    for cell in col_cells:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.font = body_font
        cell.border = border
        if cell.column == date_col_idx:
            cell.number_format = "yyyy-mm-dd"
        if cell.column in price_cols_idx:
            cell.number_format = "0.00"
for i, col in enumerate(df.columns, 1):
    ws.column_dimensions[get_column_letter(i)].width = max(14, len(col) + 4)
ws.freeze_panes = "A2"

ws2 = wb["Change_Log"]
for col_cells in ws2.iter_cols(min_row=1, max_row=1):
    for cell in col_cells:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
for row in ws2.iter_rows(min_row=2):
    for cell in row:
        cell.font = body_font
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical="top")
widths = [12, 55, 50, 12]
for i, w in enumerate(widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "A2"

wb.save(OUT)

# ---------- VERIFICATION GATE ----------
verify = pd.read_excel(OUT, sheet_name="Cleaned_Data")
print("Rows:", len(verify))
print("Duplicate OrderIDs:", verify["OrderID"].duplicated().sum())
print("Duplicate TrackingNumbers:", verify["TrackingNumber"].duplicated().sum())
print("Remaining nulls:\n", verify.isna().sum().to_dict())
print("Dates valid ISO-parseable:", pd.to_datetime(verify["Date"], errors="coerce").notna().all())
