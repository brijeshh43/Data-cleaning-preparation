
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = "../data/raw/raw_dataset.xlsx"
OUT = "../excel/Cleaned_Dataset_Excel_Formulas.xlsx"
N = 1200          
LAST = N + 1      

raw = pd.read_excel(SRC)
cols = raw.columns.tolist()  

wb = Workbook()


ws_raw = wb.active
ws_raw.title = "Raw_Data"
ws_raw.append(cols)
for r in raw.itertuples(index=False):
    ws_raw.append(list(r))


ws = wb.create_sheet("Cleaned_Data")
headers = cols + ["Duplicate_OrderID_Check", "Duplicate_Tracking_Check",
                   "CouponCode_Was_Missing", "Date_Valid_Check"]
ws.append(headers)

col_letter = {c: get_column_letter(i + 1) for i, c in enumerate(cols)}

for row in range(2, LAST + 1):
    R = lambda c: f"Raw_Data!{col_letter[c]}{row}"   
    formulas = {
        "OrderID":         f"=TRIM({R('OrderID')})",
        "Date":            f"={R('Date')}",
        "CustomerID":      f"=TRIM({R('CustomerID')})",
        "Product":         f"=PROPER(TRIM({R('Product')}))",
        "Quantity":        f"={R('Quantity')}",
        "UnitPrice":       f"=ROUND({R('UnitPrice')},2)",
        "ShippingAddress": f"=PROPER(TRIM({R('ShippingAddress')}))",
        "PaymentMethod":   f"=PROPER(TRIM({R('PaymentMethod')}))",
        "OrderStatus":     f"=PROPER(TRIM({R('OrderStatus')}))",
        "TrackingNumber":  f"=TRIM({R('TrackingNumber')})",
        "ItemsInCart":     f"={R('ItemsInCart')}",
        "CouponCode":      f'=IF(TRIM({R("CouponCode")})="","No Coupon",TRIM({R("CouponCode")}))',
        "ReferralSource":  f"=PROPER(TRIM({R('ReferralSource')}))",
        "TotalPrice":      f"=ROUND({R('TotalPrice')},2)",
    }
    for c in cols:
        ws.cell(row=row, column=cols.index(c) + 1, value=formulas[c])

    a_col = col_letter["OrderID"]
    j_col = col_letter["TrackingNumber"]
    o = len(cols) + 1
    ws.cell(row=row, column=o,
            value=f'=IF(COUNTIF($A$2:$A${LAST},A{row})>1,"DUPLICATE","Unique")')
    ws.cell(row=row, column=o + 1,
            value=f'=IF(COUNTIF($J$2:$J${LAST},J{row})>1,"DUPLICATE","Unique")')
    ws.cell(row=row, column=o + 2,
            value=f'=IF(TRIM({R("CouponCode")})="","Yes","No")')
    ws.cell(row=row, column=o + 3,
            value=f'=IF(ISNUMBER(B{row}),"Valid","INVALID")')


ws_v = wb.create_sheet("Verification_Gate")
ws_v.append(["Check", "Result", "Pass Threshold"])
ws_v.append(["Duplicate OrderID count",
             f"=SUMPRODUCT((COUNTIF(Cleaned_Data!A2:A{LAST},Cleaned_Data!A2:A{LAST})>1)*1)", 0])
ws_v.append(["Duplicate TrackingNumber count",
             f"=SUMPRODUCT((COUNTIF(Cleaned_Data!J2:J{LAST},Cleaned_Data!J2:J{LAST})>1)*1)", 0])
ws_v.append(["Invalid / non-date Date cells",
             f"=SUMPRODUCT(--NOT(ISNUMBER(Cleaned_Data!B2:B{LAST})))", 0])
ws_v.append(["Remaining blank CouponCode cells (post-fix)",
             f"=COUNTBLANK(Cleaned_Data!L2:L{LAST})", 0])
ws_v.append(["VERDICT", '=IF(SUM(B2:B5)=0,"PASS - 0% Error Rate","FAIL - Review Required")', ""])


ws_c = wb.create_sheet("Change_Log")
ws_c.append(["Change ID", "Description", "Impact", "Status"])
ws_c.append(["CR001", "Filled missing CouponCode values with explicit label 'No Coupon' (not mean/mode, since absence is meaningful, not random)",
             f'=COUNTBLANK(Raw_Data!L2:L{LAST})&" of "&{LAST-1}&" records ("&TEXT(COUNTBLANK(Raw_Data!L2:L{LAST})/{LAST-1},"0.0%")&")"', "Resolved"])
ws_c.append(["CR002", "Audited OrderID and TrackingNumber for duplicates via COUNTIF (Excel equivalent of GROUP BY ... HAVING COUNT(*) > 1)",
             '="Duplicate OrderIDs: "&Verification_Gate!B2&" | Duplicate TrackingNumbers: "&Verification_Gate!B3', "Resolved"])
ws_c.append(["CR003", "Trimmed whitespace and Proper-cased free-text fields (Product, PaymentMethod, OrderStatus, ShippingAddress, ReferralSource); ID/code columns left case-untouched",
             f'="Applied across "&{LAST-1}&" records on 5 columns"', "Resolved"])
ws_c.append(["CR004", "Applied ROUND(...,2) and a fixed 0.00 number format to UnitPrice/TotalPrice so every value displays exactly 2 decimal places (raw export had 0, 1, or 2 decimals inconsistently)",
             f'={LAST-1}&" UnitPrice and "&{LAST-1}&" TotalPrice cells standardized to 2-decimal precision"', "Resolved"])
ws_c.append(["CR005", "Validated Date column as true Excel date type and applied ISO 8601 (yyyy-mm-dd) display format",
             '="Invalid dates found: "&Verification_Gate!B4', "Resolved"])

header_fill = PatternFill("solid", start_color="2F5233", end_color="2F5233")
header_font = Font(name="Arial", bold=True, color="FFFFFF")
body_font = Font(name="Arial")
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
dup_fill = PatternFill("solid", start_color="FFC7CE", end_color="FFC7CE")

def style_header(sheet, ncols):
    for cell in sheet[1][:ncols]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    sheet.freeze_panes = "A2"

style_header(ws_raw, len(cols))
style_header(ws, len(headers))
style_header(ws_v, 3)
style_header(ws_c, 4)

date_idx = cols.index("Date") + 1
price_idx = [cols.index("UnitPrice") + 1, cols.index("TotalPrice") + 1]
dup_idx = [len(cols) + 1, len(cols) + 2]

for row in ws.iter_rows(min_row=2, max_row=LAST):
    for cell in row:
        cell.font = body_font
        cell.border = border
        if cell.column == date_idx:
            cell.number_format = "yyyy-mm-dd"
        if cell.column in price_idx:
            cell.number_format = "0.00"
        if cell.column in dup_idx and cell.value and "DUPLICATE" in str(cell.value):
            cell.fill = dup_fill

for row in ws_raw.iter_rows(min_row=2, max_row=LAST):
    for cell in row:
        cell.font = body_font
        if cell.column == date_idx:
            cell.number_format = "yyyy-mm-dd"

for sheet in (ws_v, ws_c):
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")

for i, c in enumerate(headers, 1):
    ws.column_dimensions[get_column_letter(i)].width = max(14, len(c) + 2)
for i, c in enumerate(cols, 1):
    ws_raw.column_dimensions[get_column_letter(i)].width = max(14, len(c) + 2)
ws_v.column_dimensions["A"].width = 38
ws_v.column_dimensions["B"].width = 30
ws_v.column_dimensions["C"].width = 16
ws_c.column_dimensions["A"].width = 12
ws_c.column_dimensions["B"].width = 70
ws_c.column_dimensions["C"].width = 45
ws_c.column_dimensions["D"].width = 12

wb.save(OUT)
print("Saved", OUT)
